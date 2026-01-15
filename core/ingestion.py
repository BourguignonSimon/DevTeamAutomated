from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, MutableMapping, Optional, Sequence

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency may be missing at runtime
    load_workbook = None


@dataclass
class DetectedColumns:
    text: str
    category: Optional[str]
    duration: Optional[str]


_HEADER_NORMALIZATION = re.compile(r"[^a-z0-9]+")


def _normalize_header(header: str) -> str:
    return _HEADER_NORMALIZATION.sub(" ", header.strip().lower()).strip()


def detect_useful_columns(headers: Sequence[str]) -> DetectedColumns:
    """Identify the most relevant text/category/duration columns from a table.

    The heuristic favors common label variants (e.g. "task", "description", "time (h)").
    At minimum a text column must be present, otherwise a ValueError is raised.
    """

    normalized = {header: _normalize_header(header) for header in headers}

    def _match(candidates: Iterable[str]) -> Optional[str]:
        for original, norm in normalized.items():
            for cand in candidates:
                if cand in norm:
                    return original
        return None

    text_col = _match(["task", "title", "description", "name", "summary", "request"])
    if not text_col:
        raise ValueError("no text-like column detected")

    category_col = _match(["category", "type", "classification", "bucket"])
    duration_col = _match(["duration", "time", "hours", "effort", "estimate"])

    return DetectedColumns(text=text_col, category=category_col, duration=duration_col)


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    return re.sub(r"\s+", " ", text).strip()


_CATEGORY_KEYWORDS: Mapping[str, str] = {
    "security": "security",
    "sécurité": "security",
    "securite": "security",
    "fraud": "risk",
    "finance": "finance",
    "report": "reporting",
    "rapport": "reporting",
    "data": "data quality",
    "cleanup": "data quality",
    "normalize": "data quality",
}


def infer_category(raw_category: str, task_text: str) -> str:
    cleaned = _clean_text(raw_category)
    if cleaned:
        return cleaned
    text = task_text.lower()
    for keyword, category in _CATEGORY_KEYWORDS.items():
        if keyword in text:
            return category
    return "uncategorized"


_DURATION_PATTERN = re.compile(r"(?P<num>\d+(?:\.\d+)?)[ ]*(?P<unit>[a-zA-Z]*)")
_UNIT_TO_MINUTES = {
    "m": 1,
    "min": 1,
    "mins": 1,
    "minute": 1,
    "minutes": 1,
    "h": 60,
    "hr": 60,
    "hrs": 60,
    "hour": 60,
    "hours": 60,
    "d": 8 * 60,
    "day": 8 * 60,
    "days": 8 * 60,
}


def _parse_duration_minutes(value: object) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value * 60) if value < 24 else int(value)
    text = str(value).strip()
    if not text:
        return None
    match = _DURATION_PATTERN.search(text)
    if not match:
        return None
    num = float(match.group("num"))
    unit = match.group("unit").lower()
    multiplier = _UNIT_TO_MINUTES.get(unit, 1 if unit == "" else None)
    if multiplier is None:
        return None
    minutes = int(num * multiplier)
    return minutes if minutes > 0 else None


def estimate_duration_minutes(raw_value: object, task_text: str) -> int:
    parsed = _parse_duration_minutes(raw_value)
    if parsed is not None:
        return parsed

    words = len(task_text.split())
    if words <= 1:
        return 30
    if words <= 20:
        return 60
    return 120


def normalize_rows(rows: Sequence[Mapping[str, object]]) -> List[Dict[str, object]]:
    if not rows:
        return []
    detected = detect_useful_columns(list(rows[0].keys()))

    normalized_rows: List[Dict[str, object]] = []
    for row in rows:
        text = _clean_text(row.get(detected.text))
        category_raw = row.get(detected.category) if detected.category else ""
        duration_raw = row.get(detected.duration) if detected.duration else None

        normalized_rows.append(
            {
                "title": text,
                "category": infer_category(str(category_raw) if category_raw else "", text),
                "estimated_minutes": estimate_duration_minutes(duration_raw, text),
                "source_row": dict(row),
            }
        )

    return normalized_rows


def load_csv(path: Path) -> List[MutableMapping[str, object]]:
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return [dict(row) for row in reader]


def load_excel(path: Path, sheet_name: Optional[str] = None) -> List[MutableMapping[str, object]]:
    if load_workbook is None:
        raise ImportError("openpyxl is required to read Excel files")

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows: List[MutableMapping[str, object]] = []
    header_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    try:
        header_row = next(header_iter)
    except StopIteration:
        raise ValueError("Excel sheet is empty (missing header row)") from None

    if all(h is None or str(h).strip() == "" for h in header_row):
        raise ValueError("Excel sheet is empty (missing header row)")

    headers = [str(h) if h is not None else "" for h in header_row]
    for values in ws.iter_rows(min_row=2, values_only=True):
        row: Dict[str, object] = {}
        for idx, header in enumerate(headers):
            row[header] = values[idx] if idx < len(values) else None
        rows.append(row)
    return rows
